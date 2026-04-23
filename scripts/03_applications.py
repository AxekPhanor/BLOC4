"""Produit le tableau des applications x fonctions metiers Torpier.

Sortie :
- 03-Applications.xlsx : 2 feuilles (Liste detaillee + Matrice App x Fonction)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT = Path('/home/user/BLOC4/Livrables/00-Base-Commune')

# Applications : (Nom, Editeur, Type/Hebergement, Description, Fonctions metiers couvertes, Points de vigilance)
APPLICATIONS = [
    ("Microsoft Dynamics", "Microsoft", "SaaS / Azure",
     "CRM : gestion des clients, prospects, devis, commandes",
     "Commerce | Relation client | Marketing",
     "Liaison CSV avec GesProd a remplacer"),
    ("GesProd", "Interne Torpier", "On-premise (Windows Server)",
     "Suivi des approvisionnements, stocks, production",
     "Production | Stocks | Approvisionnement | Commandes internes",
     "Outil vieillissant, lenteurs, couplage fichier"),
    ("SAGE Compta", "SAGE", "On-premise",
     "Comptabilite, gestion financiere, facturation",
     "Comptabilite | Facturation | Finance",
     "A integrer dans un ERP groupe (SDSI)"),
    ("SAGE Paie", "SAGE", "On-premise",
     "Paie et gestion RH",
     "RH | Paie | GPEC",
     "A integrer dans un ERP groupe (SDSI)"),
    ("Qualeval", "Interne Torpier", "On-premise",
     "Suivi qualite et conformite reglementaire",
     "Qualite | Production | Conformite",
     "Expose un service SOAP vers Power BI"),
    ("Microsoft Project", "Microsoft", "On-premise",
     "Planification et suivi de projets",
     "Pilotage des projets | PMO",
     "Utilise par la cellule PMO (Tiffany)"),
    ("Power BI", "Microsoft", "On-premise + service Power BI",
     "Reporting, tableaux de bord",
     "Reporting | Pilotage | Performance",
     "Sources Excel + SOAP Qualeval, a industrialiser"),
    ("PrestaShop", "PrestaShop + plugins Torpier", "SaaS / OVH",
     "E-commerce B2C/B2B, plugins Stripe",
     "Commerce | E-commerce | Relation client",
     "Extensions internes a documenter et securiser"),
    ("DocuFlow", "Interne Torpier", "On-premise (Web)",
     "GED d'entreprise",
     "Documentation | Archivage | Gouvernance",
     "A moderniser (SharePoint cible)"),
    ("Jira Service Management", "Atlassian", "SaaS / Cloud Atlassian",
     "Ticketing support utilisateurs + reclamations clients",
     "Support IT | Reclamations client | SAV",
     "Processus ITIL a renforcer (categorisation, SLA)"),
    ("ArchiCAD", "Graphisoft", "Bureau (MacBook Pro)",
     "Conception architecturale, BIM",
     "Conception | R&D | BIM",
     "Postes dedies service conception"),
    ("Revit", "Autodesk", "Bureau (PC Windows 11)",
     "Conception 3D BIM",
     "Conception | R&D | BIM",
     "Postes dedies service conception"),
    ("Microsoft Exchange", "Microsoft", "On-premise (migration en cours)",
     "Messagerie d'entreprise",
     "Messagerie | Collaboration",
     "Migration vers solution unifiee (M365)"),
    ("Active Directory (AD DS)", "Microsoft", "On-premise (Nanterre + replica Arras)",
     "Annuaire d'entreprise, authentification",
     "Securite | IAM | Gouvernance des acces",
     "Pas de replica en Finlande"),
    ("GitLab (instance)", "GitLab (interne)", "On-premise (Ubuntu Server)",
     "Gestion des sources et CI/CD",
     "Developpement | Integration",
     "Usage DSI uniquement"),
]

# Matrice App x Fonction metier
FONCTIONS = [
    "Gouvernance / COPIL",
    "Commerce & Marketing",
    "Relation client / SAV",
    "Production (structures & mobiliers)",
    "Approvisionnement & stocks",
    "Conception / R&D / BIM",
    "Qualite & Conformite",
    "Comptabilite & Facturation",
    "RH & Paie",
    "Reporting & KPI",
    "E-commerce (B2C/B2B)",
    "Documentation / GED",
    "Support IT / ITSM",
    "Securite / IAM",
    "Messagerie & Collaboration",
    "Developpement / CI/CD",
]

# couverture = { (app, fonction): True }
COUVERTURE = {
    "Microsoft Dynamics": ["Commerce & Marketing", "Relation client / SAV"],
    "GesProd": ["Production (structures & mobiliers)", "Approvisionnement & stocks"],
    "SAGE Compta": ["Comptabilite & Facturation"],
    "SAGE Paie": ["RH & Paie"],
    "Qualeval": ["Qualite & Conformite", "Production (structures & mobiliers)"],
    "Microsoft Project": ["Gouvernance / COPIL"],
    "Power BI": ["Reporting & KPI", "Gouvernance / COPIL"],
    "PrestaShop": ["E-commerce (B2C/B2B)", "Commerce & Marketing"],
    "DocuFlow": ["Documentation / GED"],
    "Jira Service Management": ["Support IT / ITSM", "Relation client / SAV"],
    "ArchiCAD": ["Conception / R&D / BIM"],
    "Revit": ["Conception / R&D / BIM"],
    "Microsoft Exchange": ["Messagerie & Collaboration"],
    "Active Directory (AD DS)": ["Securite / IAM"],
    "GitLab (instance)": ["Developpement / CI/CD"],
}

# --- Generation Excel ---
wb = Workbook()
default = wb.active
wb.remove(default)

THIN = Side(border_style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10, color="222222")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1B3A5B")
FILL_HEADER = PatternFill("solid", fgColor="1B3A5B")
FILL_ALT = PatternFill("solid", fgColor="F2F2F2")
FILL_COVERED = PatternFill("solid", fgColor="2E7D32")

# --- Feuille 1 : liste detaillee des applications ---
ws = wb.create_sheet("Applications - detail")
ws["A1"] = "Applications du SI Torpier - Liste detaillee"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:F1")
ws.row_dimensions[1].height = 30

headers = ["Application", "Editeur", "Type / hebergement",
           "Description", "Fonctions metiers couvertes",
           "Points de vigilance"]
for j, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws.row_dimensions[3].height = 30

for i, (name, editor, typ, desc, fonctions, vigilance) in enumerate(APPLICATIONS, start=4):
    for j, val in enumerate([name, editor, typ, desc, fonctions, vigilance], 1):
        c = ws.cell(row=i, column=j, value=val)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDER
        if i % 2 == 0:
            c.fill = FILL_ALT
    ws.row_dimensions[i].height = 48

widths = [25, 24, 28, 45, 38, 35]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + j)].width = w

# --- Feuille 2 : matrice App x Fonction ---
ws2 = wb.create_sheet("Matrice App x Fonction")
ws2["A1"] = "Matrice de couverture Application x Fonction metier"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(FONCTIONS) + 1)
ws2.row_dimensions[1].height = 30

# ligne d'en-tete fonctions
ws2.cell(row=3, column=1, value="Application")
c = ws2.cell(row=3, column=1)
c.font = HEADER_FONT
c.fill = FILL_HEADER
c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
c.border = BORDER
for j, fonction in enumerate(FONCTIONS, start=2):
    c = ws2.cell(row=3, column=j, value=fonction)
    c.font = HEADER_FONT
    c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True, text_rotation=60)
    c.border = BORDER
ws2.row_dimensions[3].height = 110

ws2.column_dimensions["A"].width = 26
for j in range(2, len(FONCTIONS) + 2):
    ws2.column_dimensions[chr(64 + j)].width = 8

for i, (name, *_rest) in enumerate(APPLICATIONS, start=4):
    c = ws2.cell(row=i, column=1, value=name)
    c.font = Font(name="Calibri", size=10, bold=True, color="222222")
    c.alignment = Alignment(vertical="center")
    c.border = BORDER
    covered = set(COUVERTURE.get(name, []))
    for j, fonction in enumerate(FONCTIONS, start=2):
        c = ws2.cell(row=i, column=j, value="X" if fonction in covered else "")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
        if fonction in covered:
            c.fill = FILL_COVERED
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ws2.row_dimensions[i].height = 22

# Ligne de totaux
total_row = len(APPLICATIONS) + 4
c = ws2.cell(row=total_row, column=1, value="Nombre d'applis")
c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
c.fill = FILL_HEADER
c.alignment = Alignment(horizontal="center", vertical="center")
c.border = BORDER
for j, fonction in enumerate(FONCTIONS, start=2):
    count = sum(1 for name in COUVERTURE if fonction in COUVERTURE[name])
    c = ws2.cell(row=total_row, column=j, value=count)
    c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER
ws2.row_dimensions[total_row].height = 22

xlsx_path = OUT / "03-Applications.xlsx"
wb.save(xlsx_path)
print(f"OK Excel : {xlsx_path}")
