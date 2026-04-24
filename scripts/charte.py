"""Charte graphique commune pour les livrables Torpier.

Objectifs :
- Un seul endroit pour les couleurs, polices, marges, pieds de page
- Rendu coherent et professionnel sur tous les livrables PNG
- Export 16:9 Full HD (1920 x 1080) par defaut

Usage :
    from charte import Charte, setup_figure, add_header, add_footer
"""
from __future__ import annotations

import matplotlib as mpl
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.patheffects import withStroke
from pathlib import Path

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
class Charte:
    # Palette principale (pro, sobre)
    PRIMARY = "#0A2342"      # bleu nuit profond
    PRIMARY_LIGHT = "#1E3A5F"
    SECONDARY = "#2B7A78"    # vert teal sobre
    ACCENT = "#C89B3F"       # ocre bois (clin d'oeil Torpier)
    ACCENT_LIGHT = "#F3E9D2"

    # Neutres
    INK = "#1F2937"          # texte principal
    INK_SOFT = "#4B5563"     # texte secondaire
    MUTED = "#6B7280"        # texte tertiaire
    LINE = "#E5E7EB"         # bordures fines
    BG_SOFT = "#F8FAFC"      # fond leger
    BG_CARD = "#FFFFFF"

    # Signaletique (alerte, succes, info, warn, violet)
    DANGER = "#B91C1C"
    DANGER_SOFT = "#FEE2E2"
    SUCCESS = "#047857"
    SUCCESS_SOFT = "#D1FAE5"
    INFO = "#1D4ED8"
    INFO_SOFT = "#DBEAFE"
    WARN = "#C2410C"
    WARN_SOFT = "#FFEDD5"
    VIOLET = "#6D28D9"
    VIOLET_SOFT = "#EDE9FE"

    # Typographie
    FONT = "DejaVu Sans"
    FONT_MONO = "DejaVu Sans Mono"

    # Layout (en unites figure 0..100)
    MARGIN_X = 3.5
    MARGIN_TOP = 9
    MARGIN_BOTTOM = 4

    # Projet - pied de page
    PROJECT = "Groupe Torpier | SDSI 2026-2029"
    VERSION = "v1.0"
    AUTHOR = "DSI Torpier"


# ---------------------------------------------------------------------------
# Parametres globaux matplotlib
# ---------------------------------------------------------------------------
def apply_mpl_defaults() -> None:
    mpl.rcParams.update({
        "font.family": Charte.FONT,
        "font.size": 10,
        "axes.edgecolor": Charte.LINE,
        "axes.linewidth": 0.8,
        "savefig.dpi": 180,
        "figure.dpi": 120,
        "figure.facecolor": "white",
        "savefig.facecolor": "white",
        "savefig.bbox": "tight",
    })


# ---------------------------------------------------------------------------
# Figure (canvas) standard 16:9 1920x1080 a 180dpi
# ---------------------------------------------------------------------------
def setup_figure(width: float = 100, height: float = 56.25):
    """Cree une figure 16:9 avec axes normalises 0..width, 0..height."""
    apply_mpl_defaults()
    fig = plt.figure(figsize=(19.2, 10.8))
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_xlim(0, width)
    ax.set_ylim(0, height)
    ax.set_axis_off()
    # fond leger
    ax.add_patch(patches.Rectangle((0, 0), width, height, linewidth=0,
                                   facecolor="white", zorder=-10))
    return fig, ax, width, height


# ---------------------------------------------------------------------------
# Composants : header, footer, cartes, tableaux
# ---------------------------------------------------------------------------
def add_header(ax, title: str, subtitle: str | None = None,
               width: float = 100, height: float = 56.25,
               kicker: str = "SDSI 2026-2029") -> None:
    """Ajoute un en-tete homogene en haut de la figure."""
    # bande d'accent fine
    ax.add_patch(patches.Rectangle((0, height - 0.35), width, 0.35,
                                   linewidth=0, facecolor=Charte.ACCENT,
                                   zorder=2))
    # kicker (petite etiquette au-dessus du titre)
    ax.text(Charte.MARGIN_X, height - 1.9, kicker.upper(),
            fontsize=10, fontweight="bold", color=Charte.ACCENT,
            va="center", ha="left",
            family=Charte.FONT)
    # titre principal
    ax.text(Charte.MARGIN_X, height - 3.6, title,
            fontsize=26, fontweight="bold", color=Charte.PRIMARY,
            va="center", ha="left", family=Charte.FONT)
    # sous-titre
    if subtitle:
        ax.text(Charte.MARGIN_X, height - 5.1, subtitle,
                fontsize=13, color=Charte.INK_SOFT,
                va="center", ha="left",
                family=Charte.FONT, style="italic")
    # ligne de separation
    ax.plot([Charte.MARGIN_X, width - Charte.MARGIN_X],
            [height - 6.1, height - 6.1],
            color=Charte.LINE, lw=1.2, zorder=1)


def add_footer(ax, page_label: str = "",
               width: float = 100, height: float = 56.25) -> None:
    """Ajoute un pied de page homogene."""
    ax.plot([Charte.MARGIN_X, width - Charte.MARGIN_X], [2.5, 2.5],
            color=Charte.LINE, lw=0.8, zorder=1)
    # gauche : projet
    ax.text(Charte.MARGIN_X, 1.4, Charte.PROJECT,
            fontsize=8.5, color=Charte.MUTED, va="center", ha="left",
            family=Charte.FONT)
    # centre : page label
    if page_label:
        ax.text(width / 2, 1.4, page_label,
                fontsize=8.5, color=Charte.MUTED, va="center", ha="center",
                family=Charte.FONT)
    # droite : version + auteur
    ax.text(width - Charte.MARGIN_X, 1.4,
            f"{Charte.AUTHOR}  -  {Charte.VERSION}",
            fontsize=8.5, color=Charte.MUTED, va="center", ha="right",
            family=Charte.FONT)


def shadow(ax, x, y, w, h, color=Charte.PRIMARY, alpha=0.07,
           offset=0.25, radius=0.4):
    """Simule une ombre portee avec un rectangle arrondi decale."""
    ax.add_patch(patches.FancyBboxPatch(
        (x + offset, y - offset), w, h,
        boxstyle=f"round,pad=0.05,rounding_size={radius}",
        linewidth=0, facecolor=color, alpha=alpha, zorder=1))


def card(ax, x, y, w, h, fill=Charte.BG_CARD, edge=Charte.LINE,
         lw=0.9, radius=0.4, with_shadow=True):
    """Dessine une carte rectangulaire arrondie avec ombre optionnelle."""
    if with_shadow:
        shadow(ax, x, y, w, h)
    ax.add_patch(patches.FancyBboxPatch(
        (x, y), w, h,
        boxstyle=f"round,pad=0.05,rounding_size={radius}",
        linewidth=lw, edgecolor=edge, facecolor=fill, zorder=2))


def band(ax, x, y, w, h, color=Charte.PRIMARY, radius=0.4,
         side="top"):
    """Dessine un bandeau colore (haut / gauche) d'une carte."""
    if side == "top":
        # bandeau avec coins arrondis en haut uniquement
        ax.add_patch(patches.FancyBboxPatch(
            (x, y), w, h,
            boxstyle=f"round,pad=0.05,rounding_size={radius}",
            linewidth=0, facecolor=color, zorder=3))
    else:
        ax.add_patch(patches.FancyBboxPatch(
            (x, y), w, h,
            boxstyle=f"round,pad=0.05,rounding_size={radius}",
            linewidth=0, facecolor=color, zorder=3))


def wrap(text: str, max_len: int = 28) -> str:
    """Retour a la ligne simple sur les espaces pour ne pas couper les mots."""
    out, line = [], ""
    for word in text.split():
        if len(line) + len(word) + 1 > max_len and line:
            out.append(line)
            line = word
        else:
            line = f"{line} {word}".strip()
    if line:
        out.append(line)
    return "\n".join(out)


def save(fig, path: Path) -> None:
    """Sauvegarde la figure avec les reglages par defaut et ferme."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(path, dpi=180, bbox_inches="tight", facecolor="white",
                pad_inches=0.2)
    plt.close(fig)


# ---------------------------------------------------------------------------
# Styles Excel coherents (importables par les scripts)
# ---------------------------------------------------------------------------
def excel_styles():
    """Retourne un dict de styles openpyxl coherents avec la charte."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(border_style="thin", color=Charte.LINE.replace("#", ""))
    medium = Side(border_style="medium", color=Charte.PRIMARY.replace("#", ""))

    return {
        "title": Font(name=Charte.FONT, size=18, bold=True,
                      color=Charte.PRIMARY.replace("#", "")),
        "kicker": Font(name=Charte.FONT, size=10, bold=True,
                       color=Charte.ACCENT.replace("#", "")),
        "subtitle": Font(name=Charte.FONT, size=11, italic=True,
                         color=Charte.INK_SOFT.replace("#", "")),
        "header_font": Font(name=Charte.FONT, size=11, bold=True,
                            color="FFFFFF"),
        "body_font": Font(name=Charte.FONT, size=10,
                          color=Charte.INK.replace("#", "")),
        "body_bold": Font(name=Charte.FONT, size=10, bold=True,
                          color=Charte.INK.replace("#", "")),
        "footer_font": Font(name=Charte.FONT, size=9,
                            color=Charte.MUTED.replace("#", "")),
        "fill_header": PatternFill("solid",
                                   fgColor=Charte.PRIMARY.replace("#", "")),
        "fill_accent": PatternFill("solid",
                                   fgColor=Charte.ACCENT.replace("#", "")),
        "fill_alt": PatternFill("solid",
                                fgColor="F5F7FA"),
        "fill_card": PatternFill("solid", fgColor="FFFFFF"),
        "align_center": Alignment(horizontal="center", vertical="center",
                                  wrap_text=True),
        "align_top": Alignment(horizontal="left", vertical="top",
                               wrap_text=True),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "border_medium": Border(left=medium, right=medium,
                                top=medium, bottom=medium),
    }
