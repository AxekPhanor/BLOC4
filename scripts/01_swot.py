"""Produit le SWOT Torpier - version pro (PNG) + Excel propre."""
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent))

from charte import (Charte, setup_figure, add_header, add_footer,
                    card, band, shadow, save, wrap, excel_styles)
import matplotlib.patches as patches
from openpyxl import Workbook

OUT = Path('/home/user/BLOC4/Livrables/00-Base-Commune')
OUT.mkdir(parents=True, exist_ok=True)

# =============================================================================
# Donnees
# =============================================================================
STRENGTHS = [
    "Entreprise familiale depuis 1862, notoriete etablie",
    "Croissance forte : +30 % de CA sur 2023-2025",
    "Virage e-commerce reussi (PrestaShop + bots SAV)",
    "DSI structuree (DSI, Infra, PMO, Applicatif + devs)",
    "Tiffany Valentia PMP + Scrum Master en cellule PMO",
    "Socle technique solide : SQL AlwaysOn, Veeam, AD DS",
    "Jira Service Management deploye (amorce ITIL)",
    "Direction impliquee dans la strategie SI",
    "Culture SI presente chez les utilisateurs",
]
WEAKNESSES = [
    "Aucun PRA / PCA formalise (incendie = 150 K EUR)",
    "Echanges CSV fragiles GesProd - Dynamics",
    "Commandes production via mail + Excel",
    "Pas de replica AD DS sur l'usine finlandaise",
    "VLAN unique Production + Conception (cloisonnement)",
    "Migration Exchange non finalisee",
    "Processus ITIL immatures (problemes, changements)",
    "Absence de MFA / SSO face au teletravail croissant",
    "Serveurs Windows 2019 proches de l'obsolescence",
    "Processus de pilotage des projets SI en construction",
]
OPPORTUNITIES = [
    "Expansion europeenne : DE, IT, ES, Finlande",
    "Nouveau segment B2B prive (promoteurs immobiliers)",
    "Demande clients produits ecoresponsables et traces",
    "Industrie 4.0 : IoT, maintenance predictive",
    "Maturite cloud (Microsoft 365, Azure, iPaaS)",
    "Cadres reglementaires RGPD, NIS 2, CSRD",
    "Ecosysteme d'API ouvert pour integrer les applis",
    "Freelances mobilisables pour absorber la charge",
]
THREATS = [
    "Cybermenaces amplifiees par le teletravail",
    "Obsolescence technologique (Exchange, Win 2019, VPN)",
    "Pression reglementaire RGPD, NIS 2, CSRD",
    "Dependance a des applis internes vieillissantes",
    "Penurie de competences IT et industrielles",
    "Volatilite des cours du bois et approvisionnement",
    "Concurrence europeenne accrue",
    "Risques d'incident physique (incendie, sinistre)",
]

# =============================================================================
# Generation PNG
# =============================================================================
def render_png() -> Path:
    fig, ax, W, H = setup_figure()
    add_header(
        ax,
        title="SWOT du Groupe Torpier",
        subtitle="Diagnostic strategique - synthese interne et environnement externe",
        width=W, height=H,
        kicker="Phase 1 - Analyse de l'existant",
    )

    # Zone utile pour les 4 quadrants
    zone_y_bot = 5
    zone_y_top = H - 8  # sous le header
    zone_x_left = Charte.MARGIN_X
    zone_x_right = W - Charte.MARGIN_X

    # 4 quadrants (2x2)
    gap = 1.2
    q_w = (zone_x_right - zone_x_left - gap) / 2
    q_h = (zone_y_top - zone_y_bot - gap) / 2

    quadrants = [
        # (title, items, x, y, color, soft, icon, subtitle)
        ("Forces", STRENGTHS,
         zone_x_left, zone_y_bot + q_h + gap,
         Charte.SUCCESS, Charte.SUCCESS_SOFT, "+",
         "Atouts internes a capitaliser"),
        ("Faiblesses", WEAKNESSES,
         zone_x_left + q_w + gap, zone_y_bot + q_h + gap,
         Charte.DANGER, Charte.DANGER_SOFT, "-",
         "Points faibles internes a corriger"),
        ("Opportunites", OPPORTUNITIES,
         zone_x_left, zone_y_bot,
         Charte.INFO, Charte.INFO_SOFT, ">",
         "Leviers externes a saisir"),
        ("Menaces", THREATS,
         zone_x_left + q_w + gap, zone_y_bot,
         Charte.WARN, Charte.WARN_SOFT, "!",
         "Risques externes a anticiper"),
    ]

    band_h = 2.6  # hauteur du bandeau titre dans chaque carte

    for title, items, x, y, color, soft, icon, subtitle in quadrants:
        # ombre portee
        shadow(ax, x, y, q_w, q_h)
        # carte principale (fond)
        ax.add_patch(patches.FancyBboxPatch(
            (x, y), q_w, q_h,
            boxstyle="round,pad=0.05,rounding_size=0.5",
            linewidth=0.9, edgecolor=Charte.LINE,
            facecolor="white", zorder=2))
        # bandeau de titre colore (decoupe en haut)
        ax.add_patch(patches.FancyBboxPatch(
            (x, y + q_h - band_h), q_w, band_h,
            boxstyle="round,pad=0.05,rounding_size=0.5",
            linewidth=0, facecolor=color, zorder=3))
        # masque bas du bandeau pour avoir le bandeau "collant"
        ax.add_patch(patches.Rectangle(
            (x, y + q_h - band_h), q_w, 0.5,
            linewidth=0, facecolor=color, zorder=3))
        # pastille icone
        cx, cy = x + 1.6, y + q_h - band_h / 2
        ax.add_patch(patches.Circle((cx, cy), 0.75,
                                    facecolor="white",
                                    edgecolor=color, lw=1.5, zorder=4))
        ax.text(cx, cy, icon, ha="center", va="center",
                fontsize=14, fontweight="bold", color=color,
                zorder=5, family=Charte.FONT)
        # titre
        ax.text(x + 3.2, y + q_h - band_h / 2 + 0.3, title,
                ha="left", va="center",
                fontsize=16, fontweight="bold", color="white",
                zorder=5, family=Charte.FONT)
        # sous-titre
        ax.text(x + 3.2, y + q_h - band_h / 2 - 0.7, subtitle,
                ha="left", va="center",
                fontsize=10, color="white", alpha=0.9,
                zorder=5, family=Charte.FONT, style="italic")

        # items
        item_y_top = y + q_h - band_h - 0.9
        item_y_bot = y + 0.8
        n = len(items)
        line_spacing = (item_y_top - item_y_bot) / max(n, 1)
        bullet_x = x + 1.1
        text_x = x + 1.8
        for i, item in enumerate(items):
            yy = item_y_top - i * line_spacing - 0.15
            # puce
            ax.add_patch(patches.Circle((bullet_x, yy), 0.12,
                                        facecolor=color, zorder=4))
            # texte
            ax.text(text_x, yy, item, ha="left", va="center",
                    fontsize=10.2, color=Charte.INK,
                    family=Charte.FONT, zorder=4)

    add_footer(ax, page_label="Fiche 01 - SWOT", width=W, height=H)

    out = OUT / "01-SWOT.png"
    save(fig, out)
    return out


# =============================================================================
# Generation Excel propre
# =============================================================================
def render_xlsx() -> Path:
    s = excel_styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "SWOT Torpier"

    # Page setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.sheet_view.showGridLines = False

    # Bandeau accent (ligne 1)
    ws.row_dimensions[1].height = 6

    # Kicker + titre
    ws["A2"] = "PHASE 1 - ANALYSE DE L'EXISTANT"
    ws["A2"].font = s["kicker"]
    ws["A3"] = "SWOT du Groupe Torpier"
    ws["A3"].font = s["title"]
    ws["A4"] = "Diagnostic strategique - synthese interne et environnement externe"
    ws["A4"].font = s["subtitle"]
    ws.merge_cells("A2:D2")
    ws.merge_cells("A3:D3")
    ws.merge_cells("A4:D4")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 22

    # En-tetes quadrants (ligne 6)
    from openpyxl.styles import Alignment, Font, PatternFill
    quadrants = [
        ("Forces", STRENGTHS, Charte.SUCCESS, "Atouts internes a capitaliser"),
        ("Faiblesses", WEAKNESSES, Charte.DANGER, "Points faibles a corriger"),
        ("Opportunites", OPPORTUNITIES, Charte.INFO, "Leviers externes a saisir"),
        ("Menaces", THREATS, Charte.WARN, "Risques externes a anticiper"),
    ]
    for j, (title, _items, color, subtitle) in enumerate(quadrants, 1):
        c = ws.cell(row=6, column=j, value=title.upper())
        c.font = Font(name=Charte.FONT, size=13, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=color.replace("#", ""))
        c.alignment = s["align_center"]
        c.border = s["border"]
        c2 = ws.cell(row=7, column=j, value=subtitle)
        c2.font = Font(name=Charte.FONT, size=9, italic=True, color="FFFFFF")
        c2.fill = PatternFill("solid", fgColor=color.replace("#", ""))
        c2.alignment = s["align_center"]
        c2.border = s["border"]
    ws.row_dimensions[6].height = 28
    ws.row_dimensions[7].height = 22

    max_len = max(len(q[1]) for q in quadrants)
    for j, (_title, items, _color, _subtitle) in enumerate(quadrants, 1):
        for i in range(max_len):
            cell = ws.cell(row=8 + i, column=j,
                           value=f"-  {items[i]}" if i < len(items) else "")
            cell.font = s["body_font"]
            cell.alignment = s["align_top"]
            cell.border = s["border"]
            if (8 + i) % 2 == 0:
                cell.fill = s["fill_alt"]

    for col_letter, width in zip("ABCD", [48, 48, 48, 48]):
        ws.column_dimensions[col_letter].width = width
    for r in range(8, 8 + max_len):
        ws.row_dimensions[r].height = 28

    # Footer
    footer_row = 8 + max_len + 1
    ws.cell(row=footer_row, column=1,
            value=f"{Charte.PROJECT}  |  {Charte.AUTHOR}  |  {Charte.VERSION}")
    ws.cell(row=footer_row, column=1).font = s["footer_font"]
    ws.merge_cells(start_row=footer_row, start_column=1,
                   end_row=footer_row, end_column=4)

    out = OUT / "01-SWOT.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    p1 = render_png()
    print(f"OK PNG   : {p1}")
    p2 = render_xlsx()
    print(f"OK Excel : {p2}")
