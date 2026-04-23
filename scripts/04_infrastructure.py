"""Produit le tableau des composants techniques x applications supportees.

Sortie :
- 04-Infrastructure.xlsx : 3 feuilles (Serveurs, Reseau/VLAN, Postes clients)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT = Path('/home/user/BLOC4/Livrables/00-Base-Commune')

# Serveurs : (Composant, OS, Hyperviseur, Localisation, Applications supportees, Etat / vigilance)
SERVEURS = [
    ("Serveurs Applicatifs", "Windows Server 2019/2022", "Hyper-V", "Nanterre",
     "GesProd | SAGE Compta | SAGE Paie | Qualeval | DocuFlow | MS Project",
     "Heberge les applis metier on-premise - certains OS 2019 en fin de vie"),
    ("Cluster Bases de donnees (2 noeuds)", "Windows Server 2022", "Hyper-V", "Nanterre",
     "SQL Server + moteurs secondaires (supports GesProd, SAGE, Qualeval, DocuFlow)",
     "Haute dispo AlwaysOn OK - a proteger par PRA"),
    ("Serveur Git", "Ubuntu Server 22.04", "Hyper-V", "Nanterre (VLAN 20)",
     "GitLab (versionning + CI/CD)",
     "Usage DSI, a sauvegarder"),
    ("Controleurs de domaine (AD DS)", "Windows Server 2022", "Bare-metal", "Nanterre + replica Arras",
     "Active Directory (authentification pour toutes les applis internes)",
     "Pas de replica Finlande - risque de coupure auth"),
    ("Serveur de fichiers (DFS)", "Windows Server 2022", "Hyper-V", "Nanterre",
     "Partages de fichiers (DocuFlow, dossiers services)",
     "Quotas, a migrer partiellement vers SharePoint"),
    ("Serveur de sauvegarde (Veeam)", "Windows Server 2019", "Bare-metal", "Nanterre",
     "Sauvegarde des VM et bases - support PRA",
     "PRA non documente - OS 2019 a migrer"),
    ("Serveur WSUS", "Windows Server 2019", "Hyper-V", "Nanterre",
     "Mises a jour Windows (serveurs et postes)",
     "OS a migrer"),
    ("Serveur Exchange", "Windows Server 2019", "Hyper-V", "Nanterre",
     "Messagerie d'entreprise",
     "MIGRATION EN COURS vers solution unifiee (M365)"),
    ("Serveur RDS (Bureau a distance)", "Windows Server 2019", "Hyper-V", "Nanterre",
     "Acces distant aux applis pour itinerants / teletravail",
     "A moderniser avec SSO / MFA"),
    ("Serveur Proxy et securite", "Windows Server 2019", "Bare-metal", "Nanterre",
     "Filtrage internet, securite reseau",
     "A rapprocher d'une approche Zero Trust"),
    ("Hyperviseur central", "Windows Server 2022", "Bare-metal (Hyper-V)", "Nanterre",
     "Execute les VMs du parc serveurs",
     "Socle principal de virtualisation"),
]

# Reseau / VLAN
VLANS = [
    ("VLAN 10 - Direction generale", "Acces securise pour la direction",
     "Messagerie, Power BI", "Nanterre"),
    ("VLAN 20 - DSI (postes IT)", "Postes des equipes IT + serveur de test + GitLab",
     "GitLab, outils DSI", "Nanterre"),
    ("VLAN 21 - DSI (serveurs & infra)", "Serveurs de production et equipements IT",
     "Toutes les applis internes serveurs", "Nanterre"),
    ("VLAN 30 - Administratif & RH", "Salaries des services financiers et RH",
     "SAGE Compta, SAGE Paie, DocuFlow", "Nanterre"),
    ("VLAN 40 - Commercial & Marketing", "CRM, relation client et webmarketing",
     "Microsoft Dynamics, PrestaShop (admin)", "Nanterre + bureaux EU (via VPN)"),
    ("VLAN 50 - Production & Conception", "VLAN unique pour fabrication et conception",
     "GesProd, Qualeval, ArchiCAD, Revit", "Arras (a cloisonner)"),
    ("VLAN 60 - Usine finlandaise", "Reseau distinct pour le site finlandais",
     "GesProd, Qualeval (via VPN)", "Finlande"),
]

# Equipements reseau
EQUIPEMENTS = [
    ("Switches coeur de reseau", "Cisco Catalyst", "Nanterre + Arras",
     "Interconnecte tous les VLANs"),
    ("Switches d'acces", "Cisco SG Series", "Nanterre + Arras + Finlande",
     "Distribution des postes et imprimantes"),
    ("Routeurs inter-sites", "Cisco ISR 4000 (redondance + QoS)", "Nanterre + Arras",
     "Interconnecte les sites avec QoS"),
    ("Firewall / UTM", "FortiGate 200E", "Nanterre", "Filtrage, IPS, terminaison VPN"),
    ("Solution VPN", "Fortinet IPSec (site-to-site)", "Nanterre + Arras + Finlande",
     "Interconnecte Nanterre/Arras/Finlande - bureaux EU via client VPN"),
]

# Postes clients
POSTES = [
    ("PC fixes", "Windows 11", "Tous sites", "Applis metier, messagerie, bureautique"),
    ("PC portables", "Windows 11", "Commerciaux itinerants + teletravailleurs",
     "Dynamics, RDS, VPN, messagerie"),
    ("PC fixes dedies conception", "Windows 11", "Service conception (Arras)",
     "Revit"),
    ("MacBook Pro", "macOS 11", "Service conception (Arras)", "ArchiCAD"),
    ("Imprimantes reseau", "Integrees a l'AD", "1 par service (2 conception, 3 par usine)",
     "Spooler / serveur d'impression AD"),
]

# --- Generation Excel ---
wb = Workbook()
default = wb.active
wb.remove(default)

THIN = Side(border_style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10, color="222222")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1B3A5B")
FILL_HEADER = PatternFill("solid", fgColor="1B3A5B")
FILL_ALT = PatternFill("solid", fgColor="F2F2F2")

def build_sheet(ws, title, headers, rows, widths, heights=48):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(headers))
    ws.row_dimensions[1].height = 30
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[3].height = 30
    for i, row in enumerate(rows, start=4):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER
            if i % 2 == 0:
                c.fill = FILL_ALT
        ws.row_dimensions[i].height = heights
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + j)].width = w

# Feuille 1 : Serveurs
ws1 = wb.create_sheet("Serveurs")
build_sheet(ws1,
            "Infrastructure serveurs - Applications supportees",
            ["Composant", "OS", "Hyperviseur", "Localisation",
             "Applications supportees", "Etat / Vigilance"],
            SERVEURS,
            [30, 24, 16, 22, 44, 42],
            heights=55)

# Feuille 2 : Reseau VLAN + equipements
ws2 = wb.create_sheet("Reseau & VLAN")
# Partie VLAN
ws2.cell(row=1, column=1, value="Architecture reseau - VLAN").font = TITLE_FONT
ws2.merge_cells("A1:D1")
ws2.row_dimensions[1].height = 30
headers = ["VLAN", "Description", "Applications / usages", "Localisation"]
for j, h in enumerate(headers, 1):
    c = ws2.cell(row=3, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws2.row_dimensions[3].height = 30
for i, row in enumerate(VLANS, start=4):
    for j, val in enumerate(row, 1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDER
        if i % 2 == 0:
            c.fill = FILL_ALT
    ws2.row_dimensions[i].height = 42
for j, w in enumerate([32, 48, 38, 30], 1):
    ws2.column_dimensions[chr(64 + j)].width = w

# Partie equipements
start = 4 + len(VLANS) + 2
ws2.cell(row=start, column=1, value="Equipements reseau").font = TITLE_FONT
ws2.merge_cells(start_row=start, start_column=1, end_row=start, end_column=4)
ws2.row_dimensions[start].height = 28
headers2 = ["Equipement", "Modele / technologie", "Localisation", "Role"]
for j, h in enumerate(headers2, 1):
    c = ws2.cell(row=start + 2, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
ws2.row_dimensions[start + 2].height = 28
for i, row in enumerate(EQUIPEMENTS, start=start + 3):
    for j, val in enumerate(row, 1):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDER
        if i % 2 == 0:
            c.fill = FILL_ALT
    ws2.row_dimensions[i].height = 40

# Feuille 3 : Postes clients
ws3 = wb.create_sheet("Postes clients")
build_sheet(ws3,
            "Postes clients et peripheriques",
            ["Type de poste", "OS", "Utilisateurs concernes", "Applications principales"],
            POSTES,
            [30, 22, 42, 40],
            heights=40)

xlsx_path = OUT / "04-Infrastructure.xlsx"
wb.save(xlsx_path)
print(f"OK Excel : {xlsx_path}")
