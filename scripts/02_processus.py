"""Produit la cartographie des processus metiers Torpier.

Sortie :
- 02-Processus-Metiers.xlsx : 1 feuille par famille + 1 feuille synthese
- 02-Processus-Cartographie.png : cartographie visuelle type "pilotage / metier / support"
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT = Path('/home/user/BLOC4/Livrables/00-Base-Commune')

# --- Structure des processus ---
# Famille -> liste de (Processus, Taches principales, Acteurs, Applications cles)
PROCESSUS = {
    "Pilotage (gouvernance)": [
        ("Gouvernance d'entreprise",
         "Definir la strategie | Piloter les COPIL | Arbitrer les budgets | Gerer les risques | Assurer la conformite",
         "Sophie Crelin (DG) | Luc Ciel-Marchand (DAF) | Ludovic Besnier (DSI)",
         "MS Project | Power BI | DocuFlow"),
        ("Pilotage des projets SI (en construction)",
         "Cadrage projet | Suivi du portefeuille | Pilotage de projet | Reporting COPIL SI",
         "Tiffany Valentia (PMO) | Ludovic Besnier (DSI) | Laurent Vigne",
         "MS Project | Jira | Power BI"),
    ],
    "Metiers coeur": [
        ("Gestion commerciale et relation client",
         "Prospection | Devis | Commande | Suivi vente | Livraison",
         "Stephanie Wolicz | Commerciaux seniors | Elodie Bouchet | Bureaux EU",
         "Microsoft Dynamics | PrestaShop | GesProd"),
        ("Gestion des stocks et de la livraison",
         "Suivi des stocks | Preparation commande | Expedition | Suivi transporteur",
         "Pole fabrication Arras | Sofia Jansson (Finlande) | Transporteurs externes",
         "GesProd | Dynamics"),
        ("Gestion des reclamations client",
         "Prise en charge | Traitement litige | Reponse client | Escalade juridique",
         "Antoine Zorba | Gregory Souza | Stephanie Wolicz",
         "Jira Service Management | Dynamics"),
        ("Conception et R&D",
         "Etude besoin | Conception CAO/BIM | Prototypage | Industrialisation",
         "Samir Oufrani | Jacques Noceri | Maria Tran | Madison Wilson",
         "ArchiCAD | Revit | DocuFlow"),
        ("Approvisionnement en matiere premiere",
         "Planification | Commande fournisseur | Reception | Controle qualite",
         "Frederic Szymon | Sofia Jansson | Atelier",
         "GesProd | SAGE Compta"),
        ("Fabrication (structures et mobiliers)",
         "Ordonnancement | Fabrication | Controle qualite | Mise en stock",
         "Frederic Szymon | Sofia Jansson | Artisans | Madison Wilson",
         "GesProd | Qualeval"),
    ],
    "Supports metiers": [
        ("Gestion comptable et facturation",
         "Saisie factures | Reglement fournisseurs | Facturation client | Consolidation",
         "Nathalie Roland | Marion Ferucci",
         "SAGE Compta | Power BI"),
        ("Gestion des ressources humaines",
         "Recrutement | Paie | GPEC | Formation",
         "Pierre Courtin | Nathalie Roland",
         "SAGE Paie | DocuFlow"),
        ("Reporting et pilotage de la performance",
         "Collecte KPI | Tableaux de bord | Analyses | Diffusion COPIL",
         "Luc Ciel-Marchand | Nathalie Roland | Ludovic Besnier",
         "Power BI | SAGE Compta | Qualeval"),
    ],
    "Supports IT": [
        ("Gestion des incidents et du support utilisateur",
         "Prise d'appel | Qualification | Resolution | Escalade | Cloture",
         "Service Desk DSI | Admin sys | Admin reseau",
         "Jira Service Management"),
        ("Maintenance de l'infrastructure",
         "Maintenance preventive | Mises a jour | Supervision | Sauvegardes",
         "Aurelie Duchamp | Admin reseau | Admin sys",
         "Veeam | WSUS | Hyper-V | FortiGate"),
        ("Developpement et integration logicielle",
         "Dev | Integration | Tests | Mise en production",
         "Laurent Vigne | Dev full-stack | Dev junior | Freelances",
         "GitLab | Jira | MS Project"),
        ("Gestion documentaire",
         "Depot documents | Versionnement | Recherche | Archivage",
         "Tous les services",
         "DocuFlow | SharePoint (cible)"),
        ("Securite SI et conformite (IAM, RGPD, PRA)",
         "Gestion des identites | Controle des acces | PRA | RGPD",
         "Aurelie Duchamp (RSSI relaye) | Ludovic Besnier",
         "AD DS | FortiGate | Veeam"),
    ],
}

# --- Generation Excel ---
wb = Workbook()
default = wb.active
wb.remove(default)

THIN = Side(border_style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10, color="222222")
FAMILY_FILLS = {
    "Pilotage (gouvernance)": "1B3A5B",
    "Metiers coeur": "2E7D32",
    "Supports metiers": "6A1B9A",
    "Supports IT": "EF6C00",
}

def fill(color_hex):
    return PatternFill("solid", fgColor=color_hex)

def add_sheet(family, lignes):
    ws = wb.create_sheet(family[:31])
    # Titre
    ws.cell(row=1, column=1, value=f"Processus - {family}").font = Font(
        name="Calibri", size=16, bold=True, color="1B3A5B")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 30
    # En-tetes
    headers = ["Processus", "Taches principales", "Acteurs impliques", "Applications cles"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = fill(FAMILY_FILLS[family])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[3].height = 28
    # Contenu
    for i, (proc, taches, acteurs, apps) in enumerate(lignes, start=4):
        for j, val in enumerate([proc, taches, acteurs, apps], 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = BODY_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER
        ws.row_dimensions[i].height = 65
    widths = [35, 55, 40, 35]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + j)].width = w

# Feuille de synthese en premier
synth = wb.create_sheet("Synthese")
synth.cell(row=1, column=1, value="Synthese - Processus Torpier").font = Font(
    name="Calibri", size=16, bold=True, color="1B3A5B")
synth.merge_cells("A1:C1")
synth.row_dimensions[1].height = 30
headers = ["Famille", "Nombre de processus", "Processus"]
for j, h in enumerate(headers, 1):
    c = synth.cell(row=3, column=j, value=h)
    c.font = HEADER_FONT
    c.fill = fill("1B3A5B")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
synth.row_dimensions[3].height = 26
for i, (family, lignes) in enumerate(PROCESSUS.items(), start=4):
    for j, val in enumerate([family, str(len(lignes)),
                             " | ".join(p[0] for p in lignes)], 1):
        c = synth.cell(row=i, column=j, value=val)
        c.font = BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDER
    synth.row_dimensions[i].height = 55
synth.column_dimensions["A"].width = 30
synth.column_dimensions["B"].width = 22
synth.column_dimensions["C"].width = 80

for family, lignes in PROCESSUS.items():
    add_sheet(family, lignes)

xlsx_path = OUT / "02-Processus-Metiers.xlsx"
wb.save(xlsx_path)
print(f"OK Excel : {xlsx_path}")

# --- Generation PNG : cartographie des processus ---
import matplotlib.pyplot as plt
import matplotlib.patches as patches

fig, ax = plt.subplots(figsize=(22, 13))
ax.set_xlim(0, 22)
ax.set_ylim(0, 13)
ax.axis("off")

ax.text(11, 12.4, "Cartographie des processus - Groupe Torpier",
        ha="center", fontsize=24, fontweight="bold", color="#1B3A5B")

# Fleche de chaine de valeur (au-dessus des boites, pas au milieu)
ax.annotate("", xy=(21.5, 11.6), xytext=(3.2, 11.6),
            arrowprops=dict(arrowstyle="->", lw=2.5, color="#2E7D32"))
ax.text(12.3, 11.8, "CHAINE DE VALEUR - de la demande au client",
        ha="center", fontsize=12, color="#2E7D32", fontweight="bold")

def draw_family(y, h, title, lignes, fill_color, text_color="#FFFFFF"):
    # bandeau gauche avec titre
    ax.add_patch(patches.Rectangle((0.2, y), 2.8, h, linewidth=0,
                                    facecolor=fill_color))
    ax.text(1.6, y + h / 2, title, ha="center", va="center",
            fontsize=13, fontweight="bold", color=text_color, rotation=90)
    # boites processus
    n = len(lignes)
    box_w = 18.3 / n
    for i, (proc, *_rest) in enumerate(lignes):
        x0 = 3.2 + i * box_w
        ax.add_patch(patches.FancyBboxPatch((x0 + 0.08, y + 0.15), box_w - 0.20,
                                             h - 0.30, boxstyle="round,pad=0.02",
                                             linewidth=1.2, edgecolor=fill_color,
                                             facecolor="#FFFFFF"))
        # wrap manuel si trop long
        label = proc
        if len(label) > 22:
            # coupe a l'espace le plus proche du milieu
            mid = len(label) // 2
            left = label.rfind(" ", 0, mid)
            right = label.find(" ", mid)
            cut = left if (mid - left) <= (right - mid) and left > 0 else right
            if cut > 0:
                label = label[:cut] + "\n" + label[cut + 1:]
        ax.text(x0 + box_w / 2, y + h / 2, label, ha="center", va="center",
                fontsize=10, color="#222222", fontweight="bold")

# positions (du haut vers le bas) : (famille_key, label_affiche, y, h, color)
layout = [
    ("Pilotage (gouvernance)", "PILOTAGE", 9.4, 1.7, "#1B3A5B"),
    ("Metiers coeur", "METIERS", 5.9, 3.3, "#2E7D32"),
    ("Supports metiers", "SUPPORTS METIERS", 3.5, 2.2, "#6A1B9A"),
    ("Supports IT", "SUPPORTS IT", 0.8, 2.5, "#EF6C00"),
]
for family_key, label, y, h, color in layout:
    draw_family(y, h, label, PROCESSUS[family_key], color)

png_path = OUT / "02-Processus-Cartographie.png"
plt.tight_layout()
plt.savefig(png_path, dpi=200, bbox_inches="tight", facecolor="white")
plt.close()
print(f"OK PNG   : {png_path}")
